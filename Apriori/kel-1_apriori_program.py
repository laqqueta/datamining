import openpyxl
import itertools

def unique_item(item_set):
    data = []

    for item in item_set:
        for it in item:
            if it not in data:
                data.append(it)

    return data

def conv_final_itemset(final_items):
    data = []

    for fi in final_items:
        for f in fi:
            data.append(f)
    
    return data

def apriori(path_file, min_support, total_data):
    load_file = openpyxl.load_workbook(path_file)
    sheet_obj = load_file.active

    items = []
    freq_datas = []
    freq_total = 0
    total_f_items = []
    total_f_freq = []
    
    for i in range(1, total_data+1, 1):
        for j in range(1, 9, 1):
            col_data = sheet_obj.cell(row=i, column=j).value
            
            if col_data != None:
                s = col_data.strip()
                if s not in items:
                    items.append(s)
    
    for item in items:
        for i in range(1, total_data+1, 1):
            for j in range(1, 9, 1):
                col_data = sheet_obj.cell(row=i, column=j).value
                
                if col_data != None:
                    s = col_data.strip()
                    if item == s:
                        freq_total += 1
        
        freq_datas.append((freq_total/total_data))
        freq_total = 0

    print("C-1")
    print("="*80)
    for i in range(len(items)):
        print(items[i],"=> %0.5f" % freq_datas[i])
    
    l = len(freq_datas)
    for i in range(l):
        for j in range(len(freq_datas)):
            if freq_datas[j] < min_support:
                freq_datas.pop(j)
                items.pop(j)
                break

    print("\nF-1")
    print("="*80)
    for i in range(len(items)):
        print(items[i],"=> %0.5f" % freq_datas[i])
        total_f_items.append(items[i])
        total_f_freq.append(freq_datas[i])

    k = 2
    final_item_set = []

    while len(items) > 2:
        temp_items = []
        freq_datas.clear()
        f = False

        temp_items.clear()

        if k > 2:
            tmp_itm = unique_item(items)
        else:
            tmp_itm = (item for item in items)

        for pair in itertools.combinations(tmp_itm, k):
            temp_items.append(list(pair))

        print(f"\nC-{k}")
        print("="*80)

        for item in temp_items:
            print(item, end=" ")
            for i in range(1, total_data+1, 1):
                data = []
                for j in range(1, 9, 1):
                    col_data = sheet_obj.cell(row=i, column=j).value
                    if col_data != None:
                        s = col_data.strip()
                        data.append(s)
                a = 0
                for d in data:
                    for it in item:
                        if d == it:
                            a += 1
                if a >= k:
                    freq_total += 1
                    f = True

            print("=> %0.5f" % (freq_total/total_data))
            if f:
                freq_datas.append((freq_total/total_data))
            else:
                freq_datas.append(0)
            freq_total = 0

        print(f"\nF-{k}")
        print("="*80)

        l = len(freq_datas)
        for i in range(l):
            for j in range(len(freq_datas)):
                if freq_datas[j] < min_support:
                    freq_datas.pop(j)
                    temp_items.pop(j)
                    break

        for i,item in enumerate(temp_items):
            print(item,"=> %0.5f" % freq_datas[i])
            total_f_items.append(item)
            total_f_freq.append(freq_datas[i])

        k+=1
        items = temp_items
        final_item_set = items

    load_file.close()
    return total_f_items, total_f_freq, final_item_set

def rules(item_set, freq, min_conf):
    # data_test = [['Tornadoes', 'Flooding', 'severe storms']]
    data_tmp = conv_final_itemset(item_set)

    fi_len = len(data_tmp)
    total_rule = ((fi_len-1)**fi_len) - (fi_len-1)
    data_pair = []

    n = 1

    while n <= (total_rule/fi_len):
        for pair in itertools.combinations(data_tmp, n):
            d = []
            d.extend(pair)
            data_pair.append(d)
        n+=1
        
        if n == 4:
            break

    f_temp = []

    for pd in data_pair:
        f_temp = []
        for fi in data_tmp:
            if fi not in pd:
                f_temp.extend([fi])

        if f_temp not in pd:
            pd.append(f_temp)

    data_rule = data_pair
    generated_rule_items = list([dp for dp in data_pair])

    final_rule_conf_tmp = []

    for i, dr in enumerate(data_rule):
        count_conf = freq[len(freq)-1]/freq[i]
        final_rule_conf_tmp.append(count_conf)

    generated_rule_item_conf = list([ac for ac in final_rule_conf_tmp])

    asc_len = len(data_rule)
    for i in range(asc_len):
        for j in range(len(data_rule)):
            if final_rule_conf_tmp[j] < min_conf:
                data_rule.pop(j)
                final_rule_conf_tmp.pop(j)
                break
            
    final_rule_item = data_rule
    final_rule_conf = final_rule_conf_tmp

    return generated_rule_items, generated_rule_item_conf, final_rule_item, final_rule_conf

def create_data(path_file, path_saveto):
    load_file = openpyxl.load_workbook(path_file)
    sheet_obj = load_file.active

    load_s = openpyxl.load_workbook(path_saveto)
    sheet_s = load_s.active


    col = 1
    row = 2

    print("Creating data")

    while sheet_obj.cell(row=row, column=2).value != None: 
        a = sheet_obj.cell(row=row, column=2).value
        a_data = a.split(',')

        for ad in a_data:
            sheet_s.cell(row=row,column=col).value = ad.strip()
            col += 1

        col = 1
        row += 1

    load_s.save(path_saveto)
    load_file.close()
    load_s.close()

    print("Done")

def main():
    min_support = 10/100
    min_confidence = 65/100
    total_data = 754

    # Ganti path file dengan lokasi file data.xlsx
    file = "data-uas.xlsx"
    # saveto = "E://Z//Kuliah//Semester 3//Data Mining//Tugas pertemuan 9//data-uas.xlsx"

    # create_data(file, saveto)

    f_items, f_freq, final_itemset = apriori(file, min_support, total_data)
    generated_rule_items, generated_rule_item_conf, final_rule, final_conf = rules(final_itemset, f_freq, min_confidence)

    print()
    print("="*119)
    print("Rule yang terbentuk:")

    for i, dr in enumerate(generated_rule_items):
        print(dr, end=' ')
        print("Support: %0.4f; Confidence: %0.4f/%0.4f" % (f_freq[len(f_freq)-1], f_freq[len(f_freq)-1], f_freq[i]), end=' ')
        print("= %0.4f" % generated_rule_item_conf[i])

    print(f"\nRule yang memenuhi syarat (minimum support = {min_support*100}% dan minimum confidence = {min_confidence*100}%):")

    for i, asc_item in enumerate(final_rule):
        print(asc_item, "=> %0.4f" % final_conf[i])

if __name__ == "__main__":
    main()
